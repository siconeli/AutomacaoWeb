# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support.select import Select
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import datetime
import openpyxl


driver = webdriver.Chrome()
driver.get('http://45.188.183.155:8079/transparencia/')
print('-> Carregamento de site realizado:')

nome_aba = '2023'

meses_calendario = ["janeiro de 2023", "fevereiro de 2023", "março de 2023", "abril de 2023", "maio de 2023", "junho de 2023", "julho de 2023", "agosto de 2023", "setembro de 2023", "outubro de 2023", "novembro de 2023", "dezembro de 2023"]

valores_receita = ['86.407,84', '357.501,18', '509.199,84', '672.832,23', '546.339,70', '994.110,77', '166.595,99', '555.271,64', '475.572,04']

workbook = openpyxl.load_workbook('irrf.xlsx')

try:
    aba_2023 = workbook[nome_aba]

    aba_2023['A1'].value = "Período"
    aba_2023['B1'].value = "Valor Receita"

    for index, linha in enumerate(aba_2023.iter_rows(min_row=2, max_row=len(valores_receita)+1, min_col=1, max_col=1)):
        for celula in linha:
            celula.value = meses_calendario[index]

    for index, linha in enumerate(aba_2023.iter_rows(min_row=2, max_row=len(valores_receita)+1, min_col=2, max_col=2)):
        print(index)
        for celula in linha:
            celula.value = valores_receita[index]
            print(valores_receita[index])

    workbook.save('irrf.xlsx')
    driver.close()

except Exception as error:
    workbook.create_sheet(nome_aba)

    aba_2023 = workbook[nome_aba]

    aba_2023['A1'].value = "Período"
    aba_2023['B1'].value = "Valor Receita"

    for index, linha in enumerate(aba_2023.iter_rows(min_row=2, max_row=len(valores_receita)+1, min_col=1, max_col=1)):
        for celula in linha:
            celula.value = meses_calendario[index]

    for index, linha in enumerate(aba_2023.iter_rows(min_row=2, max_row=len(valores_receita)+1, min_col=2, max_col=2)):
        for celula in linha:
            celula.value = valores_receita[index]

    workbook.save('irrf.xlsx')
    driver.close()